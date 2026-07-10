"""Build BigQuery optimization review artifacts for configured jobs.

This script copies the local SQL templates, fetches the matching BigQuery
metadata, optionally triggers the optimization workflow once, and writes the
numbered files used by the Conifer review folders.

Updated by: Ajman Mocsana
Additional Features: 
1) Added a documentation on the generated artifacts
2) Concurrently fetches the artifacts of multiple job IDs marked as "In Progress" inside the config.csv
"""

from __future__ import annotations
import json
import logging
import re
import shutil
import time
from dataclasses import asdict, dataclass
from datetime import datetime
import google.auth
from google.auth.transport.requests import AuthorizedSession
from google.cloud import bigquery
from openpyxl import load_workbook
import csv
import concurrent.futures
import sys
import copy
import argparse
from pathlib import Path
from typing import List, Dict, Any
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn
from rich.console import Console
from rich.table import Table

console = Console()

ROOT_DIR = Path(__file__).resolve().parent
TEMPLATE_DIR = ROOT_DIR / "templates" / "job_id"
STATE_FILE_NAME = ".automation_state.json"
ATTEMPT_DIR_PREFIX = "attempt_"

DEFAULT_CONFIG = "config.xlsx"
DEFAULT_ENTITY = "conifer"
DEFAULT_BQ_JOB_PROJECT = "cfrdnadevdata"
DEFAULT_TEST_SP_DATASET = "staging_framework"
DEFAULT_BQ_LOCATION = "us-central1"
DEFAULT_WORKFLOW_PROJECT = "cfr-dna-dev-project"
DEFAULT_WORKFLOW_LOCATION = "us-central1"
DEFAULT_WORKFLOW_NAME = "bq_query_optimisation_workflow"
DEFAULT_PARENT_JOBS_PROJECT = "cfr-dna-prod-project3"
LEGACY_CSV_CONFIG = "config.csv"
CSV_CONFIG_SUFFIXES = {".csv"}
EXCEL_CONFIG_SUFFIXES = {".xlsx", ".xlsm"}
REQUIRED_CONFIG_COLUMNS = ("entity", "metric_name", "job_id")
AUTOMATED_DIR_NAME = "automated"

QUERY_OPTIMIZATION_TABLE_NAME = "queries_for_optimization"
OPTIMIZATION_RESULTS_TABLE_NAME = "query_ai_optimization_results"
TABLE_QUERIES_FOR_OPTIMIZATION = (
    f"{DEFAULT_BQ_JOB_PROJECT}.{DEFAULT_TEST_SP_DATASET}.{QUERY_OPTIMIZATION_TABLE_NAME}"
)
TABLE_OPTIMIZATION_RESULTS = (
    f"{DEFAULT_BQ_JOB_PROJECT}.{DEFAULT_TEST_SP_DATASET}.{OPTIMIZATION_RESULTS_TABLE_NAME}"
)
PROD_JOBS_BY_PROJECT = (
    f"`{DEFAULT_PARENT_JOBS_PROJECT}.region-{DEFAULT_BQ_LOCATION}`"
    ".INFORMATION_SCHEMA.JOBS_BY_PROJECT"
)
REFRESH_ATTEMPT_SUFFIX = " - refreshed"
TEST_SP_FILE_NAME = "10_test_sp.sql"
LEGACY_TEST_SP_FILE_NAMES = ("10_1_create_test_sp.sql", "10_2_invoke_test_sp.sql")
TEST_SP_STEP = "test_sp"

STEP_ORDER = [
    "template",
    "queries",
    "sp_details",
    "orig_sp",
    "orig_query",
    "optimized_query",
    "exec_query",
    "exec_details",
    "opt_sp",
    "update_results",
    TEST_SP_STEP,
]
SP_ONLY_STEPS = {"sp_details", "orig_sp", "opt_sp", TEST_SP_STEP}
STANDALONE_SP_ARTIFACTS = (
    "2_sp_details.sql",
    "3_orig_sp.sql",
    "8_opt_sp.sql",
    TEST_SP_FILE_NAME,
    *LEGACY_TEST_SP_FILE_NAMES,
)

OPT_START_COMMENT = "-- START OPTIMIZED QUERY"
OPT_END_COMMENT = "-- END OPTIMIZED QUERY"
EXEC_CONTEXT_MARKER = "-- INSERT STORED PROCEDURE CONTEXT HERE"
LEGACY_ARTIFACT_NAME_PATTERN = re.compile(r"^\d+_")
ATTEMPT_DIR_NAME_PATTERN = re.compile(r"^attempt_(\d+)$")
REFRESH_ATTEMPT_DIR_NAME_PATTERN = re.compile(r"^attempt_(\d+)\s+-\s+refreshed$")
PARENT_JOB_ID_COMMENT_PATTERN = re.compile(
    r"(?im)^\s*#\s*Parent\s+job\s+id:\s*(?P<parent_job_id>\S*)\s*$"
)


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
)
log = logging.getLogger("bq_artifacts")


@dataclass(frozen=True)
class ConfigItem:
    entity: str
    metric_name: str
    parent_job_id: str
    job_id: str
    status_folder: str = AUTOMATED_DIR_NAME

    @property
    def job_root(self) -> Path:
        return ROOT_DIR / self.entity.lower() / self.metric_name / self.status_folder / self.job_id

    @property
    def target_dir(self) -> Path:
        return attempt_dir(self.job_root, 1)

    @property
    def has_parent_job_id(self) -> bool:
        return bool(self.parent_job_id.strip())


@dataclass(frozen=True)
class EntityProjects:
    bq_job_project: str
    workflow_project: str
    parent_jobs_project: str | list[str] = DEFAULT_PARENT_JOBS_PROJECT
    test_sp_dataset: str = DEFAULT_TEST_SP_DATASET


@dataclass(frozen=True)
class CloudSettings:
    bq_job_project: str
    bq_location: str
    workflow_project: str
    workflow_location: str
    workflow_name: str
    test_sp_dataset: str
    parent_jobs_project: str | list[str]

    @property
    def queries_for_optimization_table(self) -> str:
        return (
            f"{self.bq_job_project}.{self.test_sp_dataset}."
            f"{QUERY_OPTIMIZATION_TABLE_NAME}"
        )

    @property
    def optimization_results_table(self) -> str:
        return (
            f"{self.bq_job_project}.{self.test_sp_dataset}."
            f"{OPTIMIZATION_RESULTS_TABLE_NAME}"
        )

    def parent_jobs_by_project_tables(self) -> list[str]:
        projects = self.parent_jobs_project
        if isinstance(projects, str):
            projects = [projects]
        return [
            f"`{project}.region-{self.bq_location}`.INFORMATION_SCHEMA.JOBS_BY_PROJECT"
            for project in projects
        ]


ENTITY_PROJECTS = {
    "conifer": EntityProjects(
        bq_job_project=DEFAULT_BQ_JOB_PROJECT,
        workflow_project=DEFAULT_WORKFLOW_PROJECT,
    ),
    "tenet": EntityProjects(
        bq_job_project="thcdnadevdata",
        workflow_project="thc-dna-dev-project",
        parent_jobs_project=[
            "thc-dna-prod-project",
            "thcdnaproddata",
            "thcdnaprodviews",
        ],
        test_sp_dataset="staging",
    ),
    "uspi": EntityProjects(
        bq_job_project="thcdnadevdata",
        workflow_project="thc-dna-dev-project",
        parent_jobs_project="uspi-dna-prod-project",
        test_sp_dataset="staging",
    ),
}


def normalize_entity(entity: str | None) -> str:
    return (entity or DEFAULT_ENTITY).strip().casefold()


def cloud_settings_for_entity(entity: str | None, args: argparse.Namespace) -> CloudSettings:
    projects = ENTITY_PROJECTS.get(normalize_entity(entity), ENTITY_PROJECTS[DEFAULT_ENTITY])
    return CloudSettings(
        bq_job_project=getattr(args, "bq_job_project", None) or projects.bq_job_project,
        bq_location=getattr(args, "bq_location", None) or DEFAULT_BQ_LOCATION,
        workflow_project=getattr(args, "workflow_project", None) or projects.workflow_project,
        workflow_location=getattr(args, "workflow_location", None) or DEFAULT_WORKFLOW_LOCATION,
        workflow_name=getattr(args, "workflow_name", None) or DEFAULT_WORKFLOW_NAME,
        test_sp_dataset=getattr(args, "test_sp_dataset", None) or projects.test_sp_dataset,
        parent_jobs_project=getattr(args, "parent_jobs_project", None)
        or projects.parent_jobs_project,
    )


@dataclass(frozen=True)
class RoutineRef:
    project: str
    dataset: str
    routine_name: str


@dataclass(frozen=True)
class SqlStatement:
    text: str
    start: int
    end: int


@dataclass(frozen=True)
class ProcedureParam:
    name: str
    mode: str
    type_sql: str


@dataclass(frozen=True)
class TempTableStatement:
    name: str
    key: str
    text: str
    order: int


@dataclass(frozen=True)
class TableRef:
    project: str
    dataset: str
    table: str
    raw: str

    @property
    def key(self) -> tuple[str, str, str]:
        return (
            self.project.casefold(),
            self.dataset.casefold(),
            self.table.casefold(),
        )

    @property
    def sql_name(self) -> str:
        return self.raw


@dataclass(frozen=True)
class DmlTargetOccurrence:
    ref: TableRef
    start: int
    end: int
    order: int


@dataclass(frozen=True)
class DmlTargetMapping:
    ref: TableRef
    temp_name: str
    order: int


class AutomationError(RuntimeError):
    """Raised when one configured item cannot be completed."""


def step_applies_to_item(step: str, item: ConfigItem) -> bool:
    return item.has_parent_job_id or step not in SP_ONLY_STEPS


def remove_standalone_sp_artifacts(job_dir: Path) -> None:
    for file_name in STANDALONE_SP_ARTIFACTS:
        path = job_dir / file_name
        if path.exists():
            path.unlink()


class ArtifactAutomation:
    def __init__(self, args: argparse.Namespace) -> None:
        self.args = args
        self.cloud_settings = cloud_settings_for_entity(DEFAULT_ENTITY, args)
        self.bq_clients: dict[tuple[str, str], bigquery.Client] = {}
        self.manage_bq_clients = True
        self.bq: bigquery.Client | None = None

    def process_item(self, item: ConfigItem) -> None:
        self.use_item_cloud_settings(item)
        if not getattr(self.args, "dry_run", False):
            migrate_legacy_attempt(item.job_root)

        if getattr(self.args, "refresh", None):
            attempt_number, _ = latest_attempt_folder(item.job_root)
            job_dir = refresh_attempt_dir(item.job_root, attempt_number)
        else:
            attempt_number = self.resolve_attempt_number(item)
            job_dir = attempt_dir(item.job_root, attempt_number)
        if getattr(self.args, "dry_run", False):
            self.print_dry_run_plan(item, attempt_number, job_dir)
            return

        mode = "refresh" if getattr(self.args, "refresh", None) else (
            "rerun_no_workflow" if getattr(self.args, "rerun_job_id_no_workflow", None)
            else ("rerun" if self.args.rerun_job_id else "normal")
        )
        state = self.load_state(job_dir, item, attempt_number, mode)
        if self.args.restart:
            state = self.reset_state(job_dir, item, attempt_number, mode)
        elif self.args.from_step:
            state = self.rewind_state_from_step(state, self.args.from_step)
            self.save_state(job_dir, state)

        log.info(
            "Processing %s / %s / %s / %s",
            item.entity,
            item.metric_name,
            item.job_id,
            job_dir.name,
        )
        if not item.has_parent_job_id:
            remove_standalone_sp_artifacts(job_dir)
        for step in STEP_ORDER:
            if not step_applies_to_item(step, item):
                log.info("Skipping standalone-query step: %s", step)
                continue
            if step in state["completed_steps"]:
                log.info("Skipping completed step: %s", step)
                continue

            log.info("Running step: %s", step)
            updates = self.run_step(step, item, job_dir, state)
            self.complete_step(job_dir, state, step, updates)

    def use_item_cloud_settings(self, item: ConfigItem) -> None:
        settings = cloud_settings_for_entity(item.entity, self.args)
        self.cloud_settings = settings
        if getattr(self.args, "dry_run", False):
            return

        if not getattr(self, "manage_bq_clients", False):
            return

        client_key = (settings.bq_job_project, settings.bq_location)
        if client_key not in self.bq_clients:
            self.bq_clients[client_key] = bigquery.Client(
                project=settings.bq_job_project,
                location=settings.bq_location,
            )
        self.bq = self.bq_clients[client_key]
        log.info(
            "BigQuery jobs will be submitted from project=%s location=%s",
            settings.bq_job_project,
            settings.bq_location,
        )

    def resolve_attempt_number(self, item: ConfigItem) -> int:
        if self.args.rerun_job_id or getattr(self.args, "rerun_job_id_no_workflow", None):
            return next_attempt_number(item.job_root, include_legacy=True)
        if self.args.attempt:
            return self.args.attempt
        return 1

    def print_dry_run_plan(
        self,
        item: ConfigItem,
        attempt_number: int,
        job_dir: Path,
    ) -> None:
        print("")
        print(
            f"DRY RUN: {item.entity} / {item.metric_name} / {item.job_id} / "
            f"attempt_{attempt_number}"
        )
        print(f"Job root: {item.job_root}")
        print(f"Target folder: {job_dir}")
        if has_legacy_artifacts(item.job_root) and not attempt_dir(item.job_root, 1).exists():
            print("Would migrate existing root-level generated files into attempt_1 first")
        print(f"Would copy template: {TEMPLATE_DIR}")
        print(
            "Would submit all BigQuery jobs from "
            f"project={self.cloud_settings.bq_job_project} "
            f"location={self.cloud_settings.bq_location}"
        )
        if item.has_parent_job_id:
            print(
                "Would query parent job metadata from "
                f"{self.cloud_settings.parent_jobs_by_project_tables()}"
            )
        else:
            print("Would skip parent job metadata lookup; no parent_job_id was provided")
        print(f"Would query/write `{self.cloud_settings.queries_for_optimization_table}`")
        print(f"Would query/write `{self.cloud_settings.optimization_results_table}`")
        workflow_name = (
            f"projects/{self.cloud_settings.workflow_project}/"
            f"locations/{self.cloud_settings.workflow_location}/"
            f"workflows/{self.cloud_settings.workflow_name}"
        )
        if self.args.rerun_job_id:
            print("Would set is_active=TRUE, trigger workflow, and fetch the new optimized SQL")
        elif getattr(self.args, "rerun_job_id_no_workflow", None):
            print(
                "Would create the next attempt and fetch the latest existing optimized SQL "
                "without triggering the workflow"
            )
        elif getattr(self.args, "refresh", None):
            print("Would fetch the latest existing optimized SQL without triggering the workflow")
        else:
            print("Would trigger workflow once if optimized SQL is missing")
        print(f"Workflow: {workflow_name}")
        if item.has_parent_job_id:
            print(f"Would generate {TEST_SP_FILE_NAME}")
        else:
            print(f"Would skip stored procedure artifacts and {TEST_SP_FILE_NAME}")

    def run_step(
        self,
        step: str,
        item: ConfigItem,
        job_dir: Path,
        state: dict[str, Any],
    ) -> dict[str, Any]:
        if step == "template":
            self.copy_template(job_dir, overwrite=self.args.restart)
            if not item.has_parent_job_id:
                remove_standalone_sp_artifacts(job_dir)
            return {}
        if step == "queries":
            self.write_queries_file(job_dir, item)
            return {}
        if step == "sp_details":
            self.write_sp_details(job_dir, item)
            return {}
        if step == "orig_sp":
            self.write_original_sp(job_dir)
            return {}
        if step == "orig_query":
            self.write_original_query(job_dir, item)
            return {}
        if step == "optimized_query":
            created_at = self.write_optimized_query(job_dir, item, state)
            return {"optimized_result": {"created_at": created_at}}
        if step == "exec_query":
            self.write_exec_query(job_dir)
            return {}
        if step == "exec_details":
            self.write_exec_details(job_dir, state)
            return {}
        if step == "opt_sp":
            self.write_optimized_sp(job_dir)
            return {}
        if step == "update_results":
            self.write_update_results(job_dir, item, state)
            return {}
        if step == TEST_SP_STEP:
            self.write_test_sp_artifacts(job_dir)
            return {}
        raise AutomationError(f"Unknown step: {step}")

    def copy_template(self, job_dir: Path, overwrite: bool) -> None:
        if not TEMPLATE_DIR.exists():
            raise AutomationError(f"Template folder does not exist: {TEMPLATE_DIR}")
        job_dir.mkdir(parents=True, exist_ok=True)

        for source in TEMPLATE_DIR.rglob("*"):
            relative = source.relative_to(TEMPLATE_DIR)
            target = job_dir / relative
            if source.is_dir():
                target.mkdir(parents=True, exist_ok=True)
                continue
            if target.exists() and not overwrite:
                continue
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)

    def write_queries_file(self, job_dir: Path, item: ConfigItem) -> None:
        source = TEMPLATE_DIR / "1_queries.sql"
        text = read_text(source)
        text = text.replace("sample_parent_job_id", item.parent_job_id)
        text = text.replace("sample_job_id", item.job_id)
        text = text.replace(PROD_JOBS_BY_PROJECT, self.cloud_settings.parent_jobs_by_project_tables()[0])
        text = text.replace(
            TABLE_QUERIES_FOR_OPTIMIZATION,
            self.cloud_settings.queries_for_optimization_table,
        )
        text = text.replace(
            TABLE_OPTIMIZATION_RESULTS,
            self.cloud_settings.optimization_results_table,
        )
        if not item.has_parent_job_id:
            text = re.sub(
                r"(?s)# SP details:\n\n.*?\n\n# Original query:",
                (
                    "# SP details:\n\n"
                    "# No parent_job_id provided; this is a standalone query.\n\n"
                    "# Original query:"
                ),
                text,
                count=1,
            )
        write_text(job_dir / "1_queries.sql", text)

    def write_sp_details(self, job_dir: Path, item: ConfigItem) -> None:
        rows = None
        for table in self.cloud_settings.parent_jobs_by_project_tables():
            sql = f"""
                SELECT job_id, query
                FROM {table}
                WHERE job_id = @parent_job_id
                LIMIT 1
            """
            try:
                rows = self.query_rows(
                    sql,
                    [bigquery.ScalarQueryParameter("parent_job_id", "STRING", item.parent_job_id)],
                )
                if rows and rows[0].get("query"):
                    log.info("Found parent job in %s", table)
                    break
            except Exception as exc:
                log.warning("Could not query %s for parent job: %s", table, exc)
                rows = None

        if not rows or not rows[0].get("query"):
            raise AutomationError(f"No SP details query found for parent_job_id={item.parent_job_id}")
        write_text(job_dir / "2_sp_details.sql", rows[0]["query"].strip() + "\n")

    def write_original_sp(self, job_dir: Path) -> None:
        sp_details = read_text(job_dir / "2_sp_details.sql")
        routine = parse_routine_ref(sp_details)
        sql = f"""
            SELECT ddl
            FROM `{routine.project}.{routine.dataset}.INFORMATION_SCHEMA.ROUTINES`
            WHERE routine_name = @routine_name
            LIMIT 1
        """
        rows = self.query_rows(
            sql,
            [bigquery.ScalarQueryParameter("routine_name", "STRING", routine.routine_name)],
        )
        if not rows or not rows[0].get("ddl"):
            raise AutomationError(
                "No routine DDL found for "
                f"{routine.project}.{routine.dataset}.{routine.routine_name}"
            )
        write_text(job_dir / "3_orig_sp.sql", rows[0]["ddl"].strip() + "\n")

    def write_original_query(self, job_dir: Path, item: ConfigItem) -> None:
        sql = f"""
            SELECT query
            FROM `{self.cloud_settings.queries_for_optimization_table}`
            WHERE job_id = @job_id
            ORDER BY created_at DESC, updated_at DESC
            LIMIT 1
        """
        rows = self.query_rows(sql, [bigquery.ScalarQueryParameter("job_id", "STRING", item.job_id)])
        if not rows or not rows[0].get("query"):
            raise AutomationError(f"No original query found for job_id={item.job_id}")
        write_text(job_dir / "4_orig_query.sql", rows[0]["query"].strip() + "\n")

    def write_optimized_query(
        self,
        job_dir: Path,
        item: ConfigItem,
        state: dict[str, Any],
    ) -> str:
        if self.args.rerun_job_id:
            result = self.generate_new_optimized_result(job_dir, item, state)
            write_text(job_dir / "5_opt_query.sql", result["optimized_sql"].strip() + "\n")
            return stringify_timestamp(result["created_at"])

        result = self.fetch_latest_optimized_result(item.job_id)
        if getattr(self.args, "rerun_job_id_no_workflow", None):
            if result is None:
                raise AutomationError(
                    "No optimized SQL exists for this rerun. "
                    "No-workflow rerun mode does not activate or run the optimization workflow."
                )
            write_text(job_dir / "5_opt_query.sql", result["optimized_sql"].strip() + "\n")
            return stringify_timestamp(result["created_at"])

        if getattr(self.args, "refresh", None):
            if result is None:
                raise AutomationError(
                    "No optimized SQL exists for this refresh. "
                    "Refresh mode does not activate or run the optimization workflow."
                )
            write_text(job_dir / "5_opt_query.sql", result["optimized_sql"].strip() + "\n")
            return stringify_timestamp(result["created_at"])

        if result is None:
            if state.get("workflow_triggered"):
                raise AutomationError(
                    "Optimized SQL is still missing, and this job has already triggered "
                    "the workflow once. Rerun with --from-step optimized_query or --restart "
                    "after checking the workflow/result table."
                )
            self.activate_optimization_job(item.job_id)
            execution_name = self.trigger_workflow()
            state["workflow_triggered"] = True
            state["workflow_execution_name"] = execution_name
            self.save_state(job_dir, state)
            self.wait_for_workflow(execution_name)
            result = self.fetch_latest_optimized_result(item.job_id)

        if result is None:
            raise AutomationError(
                "Optimized SQL is still missing after one workflow run. "
                "You can rerun from this point with --from-step optimized_query."
            )

        write_text(job_dir / "5_opt_query.sql", result["optimized_sql"].strip() + "\n")
        return stringify_timestamp(result["created_at"])

    def generate_new_optimized_result(
        self,
        job_dir: Path,
        item: ConfigItem,
        state: dict[str, Any],
    ) -> dict[str, Any]:
        started_at = state.get("rerun_started_at")
        if state.get("workflow_triggered"):
            execution_name = state.get("workflow_execution_name")
            if execution_name:
                self.wait_for_workflow(execution_name)
            if not started_at:
                raise AutomationError(
                    "Rerun checkpoint is missing rerun_started_at; start a new rerun attempt."
                )
            result = self.fetch_latest_optimized_result_after(item.job_id, started_at)
            if result is None:
                raise AutomationError(
                    "Optimized SQL is still missing for this rerun attempt after the workflow."
                )
            return result

        started_at = self.fetch_current_bq_datetime()
        state["rerun_started_at"] = stringify_timestamp(started_at)
        self.activate_optimization_job(item.job_id)
        execution_name = self.trigger_workflow()
        state["workflow_triggered"] = True
        state["workflow_execution_name"] = execution_name
        self.save_state(job_dir, state)
        self.wait_for_workflow(execution_name)

        result = self.fetch_latest_optimized_result_after(item.job_id, started_at)
        if result is None:
            raise AutomationError(
                "Optimized SQL is still missing after the rerun workflow completed. "
                "Check the workflow logs and query_ai_optimization_results."
            )
        return result

    def fetch_current_bq_datetime(self) -> datetime:
        rows = self.query_rows('SELECT CURRENT_DATETIME("America/Chicago") AS current_datetime')
        if not rows or not rows[0].get("current_datetime"):
            raise AutomationError("Could not read the current BigQuery datetime")
        return coerce_datetime(rows[0]["current_datetime"])

    def fetch_latest_optimized_result(self, job_id: str) -> dict[str, Any] | None:
        sql = f"""
            SELECT optimized_sql, created_at
            FROM `{self.cloud_settings.optimization_results_table}`
            WHERE job_id = @job_id
              AND optimized_sql IS NOT NULL
              AND TRIM(optimized_sql) != ''
            ORDER BY created_at DESC, updated_at DESC
            LIMIT 1
        """
        rows = self.query_rows(sql, [bigquery.ScalarQueryParameter("job_id", "STRING", job_id)])
        if not rows:
            return None
        return rows[0]

    def fetch_latest_optimized_result_after(
        self,
        job_id: str,
        created_after: Any,
    ) -> dict[str, Any] | None:
        sql = f"""
            SELECT optimized_sql, created_at
            FROM `{self.cloud_settings.optimization_results_table}`
            WHERE job_id = @job_id
              AND created_at > @created_after
              AND optimized_sql IS NOT NULL
              AND TRIM(optimized_sql) != ''
            ORDER BY created_at DESC, updated_at DESC
            LIMIT 1
        """
        rows = self.query_rows(
            sql,
            [
                bigquery.ScalarQueryParameter("job_id", "STRING", job_id),
                bigquery.ScalarQueryParameter(
                    "created_after",
                    "DATETIME",
                    coerce_datetime(created_after),
                ),
            ],
        )
        if not rows:
            return None
        return rows[0]

    def activate_optimization_job(self, job_id: str) -> None:
        sql = f"""
            UPDATE `{self.cloud_settings.queries_for_optimization_table}`
            SET
              is_active = TRUE,
              updated_at = CURRENT_DATETIME("America/Chicago")
            WHERE job_id = @job_id
        """
        self.query_rows(sql, [bigquery.ScalarQueryParameter("job_id", "STRING", job_id)])

    def write_exec_query(self, job_dir: Path) -> None:
        template = normalize_newlines(read_text(TEMPLATE_DIR / "6_exec_query.sql"))
        sp_details = read_text_if_exists(job_dir / "2_sp_details.sql")
        original_sp = read_text_if_exists(job_dir / "3_orig_sp.sql")
        original_query_text = read_text(job_dir / "4_orig_query.sql")
        optimized_query_text = read_text(job_dir / "5_opt_query.sql")
        context_source_sql = original_sp or original_query_text
        original_query = strip_query_destination(original_query_text)
        optimized_query = strip_query_destination(optimized_query_text)
        context_block = build_exec_query_context(
            sp_details,
            context_source_sql,
            original_query_text,
            optimized_query_text,
        )

        original_block = (
            "CREATE OR REPLACE TEMP TABLE V_TEMP_TABLE_ORIG AS\n"
            f"{ensure_semicolon(original_query)}"
        )
        optimized_block = (
            "CREATE OR REPLACE TEMP TABLE V_TEMP_TABLE_OPT AS\n"
            f"{ensure_semicolon(optimized_query)}"
        )

        original_marker = (
            "--CREATE OR REPLACE TEMP TABLE V_TEMP_TABLE_ORIG AS\n"
            "-- INSERT YOUR ORIGINAL SCRIPT HERE"
        )
        optimized_marker = (
            "--CREATE OR REPLACE TEMP TABLE V_TEMP_TABLE_OPT AS\n"
            "-- INSERT YOUR OPTIMIZED SCRIPT HERE"
        )
        if (
            EXEC_CONTEXT_MARKER not in template
            or original_marker not in template
            or optimized_marker not in template
        ):
            raise AutomationError("6_exec_query.sql template markers were not found")

        text = template.replace(EXEC_CONTEXT_MARKER, context_block.rstrip())
        text = text.replace(original_marker, original_block)
        text = text.replace(optimized_marker, optimized_block)
        write_text(job_dir / "6_exec_query.sql", text)

    def write_exec_details(self, job_dir: Path, state: dict[str, Any]) -> None:
        created_at = get_optimized_created_at(state)
        template = read_text(TEMPLATE_DIR / "7_exec_details.txt")
        text = re.sub(
            r"(?m)^OPT_CREATED:[^\r\n]*(?:\r?\n)?",
            f"OPT_CREATED: {created_at}\n",
            template,
            count=1,
        )
        write_text(job_dir / "7_exec_details.txt", text)

    def write_optimized_sp(self, job_dir: Path) -> None:
        original_sp = normalize_newlines(read_text(job_dir / "3_orig_sp.sql"))
        original_query = normalize_newlines(read_text(job_dir / "4_orig_query.sql")).strip()
        optimized_query = normalize_newlines(read_text(job_dir / "5_opt_query.sql")).strip()

        replacement = (
            f"{OPT_START_COMMENT}\n"
            f"{ensure_semicolon(optimized_query)}\n"
            f"{OPT_END_COMMENT}"
        )
        text = replace_original_query(original_sp, original_query, replacement)
        write_text(job_dir / "8_opt_sp.sql", text)

    def write_update_results(
        self,
        job_dir: Path,
        item: ConfigItem,
        state: dict[str, Any],
    ) -> None:
        created_at = get_optimized_created_at(state)
        optimized_query = normalize_newlines(read_text(job_dir / "5_opt_query.sql")).strip()
        text = (
            f"UPDATE `{self.cloud_settings.optimization_results_table}`\n"
            "SET\n"
            f'  optimized_sql = """{escape_triple_quoted_string(optimized_query)}""",\n'
            '  updated_at = CURRENT_DATETIME("America/Chicago")\n'
            f"WHERE job_id = '{escape_sql_string(item.job_id)}'\n"
            f'  AND created_at = "{escape_sql_string(created_at)}";\n'
        )
        write_text(job_dir / "9_update_results.sql", text)

    def write_test_sp_artifacts(self, job_dir: Path) -> None:
        sp_details = normalize_newlines(read_text(job_dir / "2_sp_details.sql")).strip()
        original_sp = normalize_newlines(read_text(job_dir / "3_orig_sp.sql"))
        optimized_sp = normalize_newlines(read_text(job_dir / "8_opt_sp.sql")).strip()
        original_routine = parse_create_routine_ref(original_sp)
        test_routine = RoutineRef(
            project=self.cloud_settings.bq_job_project,
            dataset=self.cloud_settings.test_sp_dataset,
            routine_name=opt_prefixed_name(original_routine.routine_name),
        )

        test_sql = build_test_sp_sql(
            optimized_sp,
            sp_details,
            test_routine,
            job_id=job_id_from_artifact_dir(job_dir),
        )
        write_text(job_dir / TEST_SP_FILE_NAME, test_sql)
        for file_name in LEGACY_TEST_SP_FILE_NAMES:
            legacy_path = job_dir / file_name
            if legacy_path.exists():
                legacy_path.unlink()

    def query_rows(
        self,
        sql: str,
        parameters: list[bigquery.ScalarQueryParameter] | None = None,
    ) -> list[dict[str, Any]]:
        if self.bq is None:
            raise AutomationError("BigQuery client is not initialized")
        job_config = bigquery.QueryJobConfig(query_parameters=parameters or [])
        log.info("Submitting BigQuery job from project=%s", self.cloud_settings.bq_job_project)
        return [dict(row) for row in self.bq.query(sql, job_config=job_config).result()]

    def trigger_workflow(self) -> str:
        session = make_authorized_session()
        parent = (
            f"projects/{self.cloud_settings.workflow_project}/"
            f"locations/{self.cloud_settings.workflow_location}/"
            f"workflows/{self.cloud_settings.workflow_name}"
        )
        url = f"https://workflowexecutions.googleapis.com/v1/{parent}/executions"
        log.info("Triggering workflow: %s", parent)
        response = session.post(url, json={})
        if response.status_code >= 400:
            raise AutomationError(
                f"Workflow trigger failed ({response.status_code}): {response.text}"
            )
        payload = response.json()
        execution_name = payload.get("name")
        if not execution_name:
            raise AutomationError(f"Workflow trigger response did not include name: {payload}")
        return execution_name

    def wait_for_workflow(self, execution_name: str) -> None:
        session = make_authorized_session()
        url = f"https://workflowexecutions.googleapis.com/v1/{execution_name}"
        deadline = time.monotonic() + getattr(self.args, "workflow_timeout_seconds", 7200)
        while True:
            response = session.get(url)
            if response.status_code >= 400:
                raise AutomationError(
                    f"Workflow status check failed ({response.status_code}): {response.text}"
                )
            payload = response.json()
            state = payload.get("state")
            log.info("Workflow state: %s", state)
            if state == "SUCCEEDED":
                return
            if state in {"FAILED", "CANCELLED"}:
                raise AutomationError(f"Workflow ended with state={state}: {payload}")
            if time.monotonic() >= deadline:
                raise AutomationError(
                    f"Timed out waiting for workflow execution: {execution_name}"
                )
            time.sleep(getattr(self.args, "workflow_poll_interval_seconds", 30))

    def load_state(
        self,
        job_dir: Path,
        item: ConfigItem,
        attempt_number: int,
        mode: str,
    ) -> dict[str, Any]:
        state_path = job_dir / STATE_FILE_NAME
        if not state_path.exists():
            return new_state(item, attempt_number, mode)
        state = json.loads(read_text(state_path))
        state.setdefault("completed_steps", [])
        state.setdefault("workflow_triggered", False)
        state.setdefault("attempt_number", attempt_number)
        state.setdefault("mode", mode)
        if state.get("config") != asdict(item):
            log.warning("Checkpoint config differs from current config for %s", job_dir)
        return state

    def reset_state(
        self,
        job_dir: Path,
        item: ConfigItem,
        attempt_number: int,
        mode: str,
    ) -> dict[str, Any]:
        state_path = job_dir / STATE_FILE_NAME
        if state_path.exists():
            state_path.unlink()
        return new_state(item, attempt_number, mode)

    def rewind_state_from_step(self, state: dict[str, Any], step: str) -> dict[str, Any]:
        step_index = STEP_ORDER.index(step)
        state["completed_steps"] = [
            existing
            for existing in state.get("completed_steps", [])
            if STEP_ORDER.index(existing) < step_index
        ]
        if step_index <= STEP_ORDER.index("optimized_query"):
            state.pop("optimized_result", None)
        return state

    def complete_step(
        self,
        job_dir: Path,
        state: dict[str, Any],
        step: str,
        updates: dict[str, Any] | None,
    ) -> None:
        if updates:
            state.update(updates)
        if step not in state["completed_steps"]:
            state["completed_steps"].append(step)
        self.save_state(job_dir, state)

    def save_state(self, job_dir: Path, state: dict[str, Any]) -> None:
        job_dir.mkdir(parents=True, exist_ok=True)
        state["updated_at"] = datetime.now().isoformat(timespec="seconds")
        write_text(job_dir / STATE_FILE_NAME, json.dumps(state, indent=2) + "\n")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate BigQuery optimization artifact folders from a config file or one item."
    )
    parser.add_argument(
        "--config",
        default=DEFAULT_CONFIG,
        help="Path to config .xlsx/.xlsm or .csv file. Defaults to config.xlsx.",
    )
    parser.add_argument("--single", action="store_true")
    parser.add_argument("--entity")
    parser.add_argument("--metric-name")
    parser.add_argument("--parent-job-id")
    parser.add_argument("--job-id")
    parser.add_argument(
        "--rerun-job-id",
        help=(
            "Create the next attempt for this job_id, reactivate it, trigger the "
            "optimization workflow, and use the newly inserted optimized SQL."
        ),
    )
    parser.add_argument(
        "--rerun-job-id-no-workflow",
        help=(
            "Create the next attempt for this job_id using the latest existing "
            "optimized SQL without running the optimization workflow."
        ),
    )
    parser.add_argument(
        "--refresh",
        help=(
            "Regenerate all artifacts for this job_id in attempt_N - refreshed "
            "using the latest existing optimized SQL without running the workflow."
        ),
    )
    parser.add_argument(
        "--attempt",
        type=positive_int,
        help="Run or resume a specific attempt number. Reruns pick the next attempt automatically.",
    )
    parser.add_argument("--restart", action="store_true")
    parser.add_argument("--from-step", choices=STEP_ORDER)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--bq-job-project",
        help="Override the entity default BigQuery data/job project.",
    )
    parser.add_argument("--bq-location", default=DEFAULT_BQ_LOCATION)
    parser.add_argument(
        "--workflow-project",
        help="Override the entity default workflow project.",
    )
    parser.add_argument("--workflow-location", default=DEFAULT_WORKFLOW_LOCATION)
    parser.add_argument("--workflow-name", default=DEFAULT_WORKFLOW_NAME)
    parser.add_argument(
        "--test-sp-dataset",
        help="Override the entity default dataset for staging tables and generated test SPs.",
    )
    parser.add_argument(
        "--parent-jobs-project",
        help=(
            "Project that hosts the parent job INFORMATION_SCHEMA.JOBS_BY_PROJECT "
            f"view. Defaults to {DEFAULT_PARENT_JOBS_PROJECT}."
        ),
    )
    parser.add_argument("--workflow-poll-interval-seconds", type=int, default=30)
    parser.add_argument("--workflow-timeout-seconds", type=int, default=7200)
    args = parser.parse_args(argv)
    if args.rerun_job_id and args.attempt:
        parser.error("--attempt cannot be used with --rerun-job-id; reruns create the next attempt")
    if args.rerun_job_id_no_workflow and args.attempt:
        parser.error(
            "--attempt cannot be used with --rerun-job-id-no-workflow; "
            "reruns create the next attempt"
        )
    if args.refresh and args.rerun_job_id:
        parser.error("--refresh cannot be used with --rerun-job-id")
    if args.refresh and args.rerun_job_id_no_workflow:
        parser.error("--refresh cannot be used with --rerun-job-id-no-workflow")
    if args.rerun_job_id and args.rerun_job_id_no_workflow:
        parser.error("--rerun-job-id cannot be used with --rerun-job-id-no-workflow")
    if args.refresh and args.attempt:
        parser.error("--attempt cannot be used with --refresh")
    if args.refresh and args.from_step:
        parser.error("--from-step cannot be used with --refresh")
    return args


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    try:
        items = load_config_items(args)
    except AutomationError as exc:
        log.error("%s", exc)
        return 2

    automation = ArtifactAutomation(args)
    failures: list[tuple[ConfigItem, Exception]] = []
    for item in items:
        try:
            automation.process_item(item)
        except Exception as exc:  # noqa: BLE001 - bulk processing should continue.
            failures.append((item, exc))
            log.error("Failed job_id=%s: %s", item.job_id, exc, exc_info=True)

    if failures:
        log.error("%d item(s) failed.", len(failures))
        for item, exc in failures:
            log.error("  %s / %s / %s: %s", item.entity, item.metric_name, item.job_id, exc)
        return 1

    log.info("Completed %d item(s).", len(items))
    return 0


def load_config_items(args: argparse.Namespace) -> list[ConfigItem]:
    has_inline_item = any([args.entity, args.metric_name, args.parent_job_id, args.job_id])
    if getattr(args, "refresh", None):
        return [load_refresh_item(args, has_inline_item)]

    if args.rerun_job_id_no_workflow:
        return [load_rerun_no_workflow_item(args, has_inline_item)]

    if args.rerun_job_id:
        return [load_rerun_item(args, has_inline_item)]

    if args.single or has_inline_item:
        return [load_single_item(args)]

    config_path = resolve_config_path(args.config)
    if config_path.exists():
        return read_config(config_path)

    log.info("No %s found; prompting for a single item.", config_path)
    return [prompt_for_item(args)]


def load_rerun_item(args: argparse.Namespace, has_inline_item: bool) -> ConfigItem:
    if has_inline_item:
        item = load_single_item(args)
        if item.job_id != args.rerun_job_id:
            raise AutomationError(
                "--rerun-job-id does not match the inline --job-id: "
                f"{args.rerun_job_id} != {item.job_id}"
            )
        return item

    return load_item_from_job_id(args.rerun_job_id, "--rerun-job-id")


def load_rerun_no_workflow_item(args: argparse.Namespace, has_inline_item: bool) -> ConfigItem:
    if has_inline_item:
        item = load_single_item(args)
        if item.job_id != args.rerun_job_id_no_workflow:
            raise AutomationError(
                "--rerun-job-id-no-workflow does not match the inline --job-id: "
                f"{args.rerun_job_id_no_workflow} != {item.job_id}"
            )
        return item

    return load_item_from_job_id(args.rerun_job_id_no_workflow, "--rerun-job-id-no-workflow")


def load_refresh_item(args: argparse.Namespace, has_inline_item: bool) -> ConfigItem:
    if has_inline_item:
        item = load_single_item(args)
        if item.job_id != args.refresh:
            raise AutomationError(
                "--refresh does not match the inline --job-id: "
                f"{args.refresh} != {item.job_id}"
            )
        return item

    return load_item_from_job_id(args.refresh, "--refresh")


def load_rerun_item_from_automated(job_id: str) -> ConfigItem:
    return load_item_from_job_id(job_id, "--rerun-job-id")


def load_item_from_job_id(job_id: str, flag_name: str) -> ConfigItem:
    matches = find_job_dirs(job_id)
    if not matches:
        search_pattern = ROOT_DIR / "<project>" / "<metric>" / "<folder>" / job_id
        raise AutomationError(
            f"No job folder found for {flag_name}={job_id}; "
            f"searched {search_pattern}"
        )
    if len(matches) > 1:
        paths = ", ".join(str(path) for path in matches)
        raise AutomationError(
            f"Multiple job folders found for {flag_name}={job_id}: {paths}. "
            "Use inline arguments to disambiguate."
        )
    return config_item_from_job_dir(matches[0])


def find_automated_job_dirs(job_id: str, root_dir: Path | None = None) -> list[Path]:
    return find_job_dirs(job_id, root_dir)


def find_job_dirs(job_id: str, root_dir: Path | None = None) -> list[Path]:
    root = ROOT_DIR if root_dir is None else root_dir
    if not root.exists():
        return []

    matches: list[Path] = []
    project_dirs = sorted(
        (child for child in root.iterdir() if child.is_dir()),
        key=lambda path: path.name.casefold(),
    )
    for project_dir in project_dirs:
        metric_dirs = sorted(
            (child for child in project_dir.iterdir() if child.is_dir()),
            key=lambda path: path.name.casefold(),
        )
        for metric_dir in metric_dirs:
            status_dirs = sorted(
                (child for child in metric_dir.iterdir() if child.is_dir()),
                key=lambda path: path.name.casefold(),
            )
            for status_dir in status_dirs:
                job_root = status_dir / job_id
                if job_root.is_dir():
                    matches.append(job_root)
    return matches


def config_item_from_automated_job_dir(job_root: Path) -> ConfigItem:
    return config_item_from_job_dir(job_root)


def config_item_from_job_dir(job_root: Path) -> ConfigItem:
    status_dir = job_root.parent
    metric_dir = status_dir.parent
    project_dir = metric_dir.parent
    parent_job_id = find_parent_job_id_for_job_root(job_root)
    return ConfigItem(
        entity=project_dir.name,
        metric_name=metric_dir.name,
        parent_job_id=parent_job_id,
        job_id=job_root.name,
        status_folder=status_dir.name,
    )


def find_parent_job_id_for_job_root(job_root: Path) -> str:
    parent_job_id = find_parent_job_id_in_state(job_root)
    if parent_job_id is not None:
        return parent_job_id

    parent_job_id = find_parent_job_id_in_queries_file(job_root)
    if parent_job_id is not None:
        return parent_job_id

    raise AutomationError(
        f"Could not determine whether {job_root} belongs to a parent stored procedure; "
        f"expected {STATE_FILE_NAME} or a Parent job id comment in 1_queries.sql."
    )


def find_parent_job_id_in_state(job_root: Path) -> str | None:
    for artifact_dir in artifact_dirs_for_job_root(job_root):
        state_path = artifact_dir / STATE_FILE_NAME
        if not state_path.exists():
            continue
        try:
            state = json.loads(read_text(state_path))
        except (OSError, json.JSONDecodeError) as exc:
            raise AutomationError(f"{state_path} could not be read: {exc}") from exc

        config = state.get("config")
        if not isinstance(config, dict):
            continue
        state_job_id = str(config.get("job_id") or "").strip()
        if state_job_id and state_job_id != job_root.name:
            continue
        if "parent_job_id" in config:
            return str(config.get("parent_job_id") or "").strip()
    return None


def find_parent_job_id_in_queries_file(job_root: Path) -> str | None:
    for artifact_dir in artifact_dirs_for_job_root(job_root):
        queries_path = artifact_dir / "1_queries.sql"
        if not queries_path.exists():
            continue
        match = PARENT_JOB_ID_COMMENT_PATTERN.search(read_text(queries_path))
        if match:
            return match.group("parent_job_id").strip()
    return None


def artifact_dirs_for_job_root(job_root: Path) -> list[Path]:
    dirs = [
        path
        for _, path in sorted(
            list_attempt_folders(job_root),
            key=lambda item: (item[0], item[1].name.casefold()),
            reverse=True,
        )
    ]
    if has_legacy_artifacts(job_root):
        dirs.append(job_root)
    return dirs


def load_single_item(args: argparse.Namespace) -> ConfigItem:
    if all([args.entity, args.metric_name, args.job_id]):
        return make_config_item(
            {
                "entity": args.entity,
                "metric_name": args.metric_name,
                "parent_job_id": args.parent_job_id or "",
                "job_id": args.job_id,
            }
        )
    return prompt_for_item(args)


def prompt_for_item(args: argparse.Namespace) -> ConfigItem:
    values = {
        "entity": args.entity or input("entity: ").strip(),
        "metric_name": args.metric_name or input("metric_name: ").strip(),
        "parent_job_id": (
            args.parent_job_id
            if args.parent_job_id is not None
            else input("parent_job_id (optional; press Enter for standalone query): ").strip()
        ),
        "job_id": args.job_id or input("job_id: ").strip(),
    }
    return make_config_item(values)


def resolve_config_path(config_arg: str) -> Path:
    path = Path(config_arg)
    if not path.is_absolute():
        path = ROOT_DIR / path
    if path.exists() or config_arg != DEFAULT_CONFIG:
        return path

    legacy_path = ROOT_DIR / LEGACY_CSV_CONFIG
    if legacy_path.exists():
        log.info("No %s found; using %s instead.", path.name, legacy_path.name)
        return legacy_path
    return path


def read_config(path: Path) -> list[ConfigItem]:
    suffix = path.suffix.lower()
    if suffix in EXCEL_CONFIG_SUFFIXES:
        return read_config_excel(path)
    if suffix in CSV_CONFIG_SUFFIXES:
        return read_config_csv(path)
    if suffix == ".xls":
        raise AutomationError(f"{path} uses legacy .xls format; save it as .xlsx")
    raise AutomationError(
        f"{path} has unsupported config file extension; use .xlsx, .xlsm, or .csv"
    )


def read_config_csv(path: Path) -> list[ConfigItem]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        items = make_config_items(path, reader.fieldnames, reader)
    if not items:
        raise AutomationError(f"{path} does not contain any config rows")
    return items


def read_config_excel(path: Path) -> list[ConfigItem]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl raises several file-specific errors.
        raise AutomationError(f"{path} could not be read as an Excel workbook: {exc}") from exc

    try:
        worksheet = next(
            (sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"),
            None,
        )
        if worksheet is None:
            raise AutomationError(f"{path} does not contain a visible worksheet")

        rows = [
            [cell_value_to_string(value) for value in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
        non_empty_rows = [row for row in rows if not is_blank_values(row)]
        if not non_empty_rows:
            raise AutomationError(f"{path} does not contain a header row")

        fieldnames = trim_trailing_empty(non_empty_rows[0])
        row_dicts = excel_row_dicts(fieldnames, non_empty_rows[1:])
        items = make_config_items(path, fieldnames, row_dicts)
        if not items:
            raise AutomationError(f"{path} does not contain any config rows")
        return items
    finally:
        workbook.close()


def make_config_items(
    path: Path,
    fieldnames: list[str] | None,
    rows: Any,
) -> list[ConfigItem]:
    if not fieldnames:
        raise AutomationError(f"{path} does not contain a header row")

    normalized_fieldnames = [str(field).strip().lower() for field in fieldnames]
    duplicate_fields = find_duplicates(normalized_fieldnames)
    if duplicate_fields:
        raise AutomationError(
            f"{path} contains duplicate column(s) after lowercasing: "
            f"{', '.join(sorted(duplicate_fields))}"
        )

    missing = set(REQUIRED_CONFIG_COLUMNS) - set(normalized_fieldnames)
    if missing:
        raise AutomationError(
            f"{path} is missing required column(s): {', '.join(sorted(missing))}"
        )

    items = []
    for row in rows:
        normalized_row = {
            str(key).strip().lower(): value
            for key, value in row.items()
            if key is not None
        }
        if is_blank_row(normalized_row):
            continue
        items.append(make_config_item(normalized_row))
    return items


def excel_row_dicts(fieldnames: list[str], rows: list[list[str]]) -> list[dict[str, str]]:
    row_dicts = []
    for row in rows:
        row_dicts.append(
            {
                fieldnames[index]: row[index] if index < len(row) else ""
                for index in range(len(fieldnames))
            }
        )
    return row_dicts


def trim_trailing_empty(values: list[str]) -> list[str]:
    trimmed = list(values)
    while trimmed and not str(trimmed[-1] or "").strip():
        trimmed.pop()
    return trimmed


def is_blank_values(values: list[Any]) -> bool:
    return all(not str(value or "").strip() for value in values)


def is_blank_row(row: dict[str, Any]) -> bool:
    return all(not str(value or "").strip() for value in row.values())


def cell_value_to_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def make_config_item(row: dict[str, Any]) -> ConfigItem:
    values = {
        "entity": clean_required(row.get("entity"), "entity"),
        "metric_name": clean_required(row.get("metric_name"), "metric_name"),
        "parent_job_id": clean_optional(row.get("parent_job_id")),
        "job_id": clean_required(row.get("job_id"), "job_id"),
    }
    return ConfigItem(**values)


def clean_required(value: Any, field_name: str) -> str:
    cleaned = str(value or "").strip()
    if not cleaned:
        raise AutomationError(f"Missing required value: {field_name}")
    return cleaned


def clean_optional(value: Any) -> str:
    return str(value or "").strip()


def positive_int(value: str) -> int:
    parsed = int(value)
    if parsed < 1:
        raise argparse.ArgumentTypeError("must be greater than or equal to 1")
    return parsed


def new_state(item: ConfigItem, attempt_number: int, mode: str) -> dict[str, Any]:
    return {
        "version": 2,
        "config": asdict(item),
        "attempt_number": attempt_number,
        "mode": mode,
        "completed_steps": [],
        "workflow_triggered": False,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }


def attempt_dir(job_root: Path, attempt_number: int) -> Path:
    return job_root / f"{ATTEMPT_DIR_PREFIX}{attempt_number}"


def refresh_attempt_dir(job_root: Path, attempt_number: int) -> Path:
    return job_root / f"{ATTEMPT_DIR_PREFIX}{attempt_number}{REFRESH_ATTEMPT_SUFFIX}"


def list_attempt_numbers(job_root: Path, include_legacy: bool = False) -> list[int]:
    numbers = {number for number, _ in list_attempt_folders(job_root)}

    if include_legacy and has_legacy_artifacts(job_root) and 1 not in numbers:
        numbers.add(1)
    return sorted(numbers)


def next_attempt_number(job_root: Path, include_legacy: bool = False) -> int:
    numbers = list_attempt_numbers(job_root, include_legacy=include_legacy)
    return max(numbers, default=0) + 1


def list_attempt_folders(job_root: Path) -> list[tuple[int, Path]]:
    if not job_root.exists():
        return []

    folders: list[tuple[int, Path]] = []
    for child in job_root.iterdir():
        if not child.is_dir():
            continue
        number = parse_attempt_dir_number(child.name)
        if number is not None:
            folders.append((number, child))
    return sorted(folders, key=lambda item: (item[0], item[1].name.casefold()))


def latest_attempt_folder(job_root: Path) -> tuple[int, Path]:
    folders = list_attempt_folders(job_root)
    if not folders:
        raise AutomationError(f"No existing attempt folder found under {job_root}")
    return max(folders, key=lambda item: (item[0], item[1].name.casefold()))


def parse_attempt_dir_number(name: str) -> int | None:
    match = ATTEMPT_DIR_NAME_PATTERN.fullmatch(name)
    if not match:
        match = REFRESH_ATTEMPT_DIR_NAME_PATTERN.fullmatch(name)
    if not match:
        return None
    return int(match.group(1))


def has_legacy_artifacts(job_root: Path) -> bool:
    if not job_root.exists():
        return False
    return any(is_legacy_artifact(child) for child in job_root.iterdir())


def is_legacy_artifact(path: Path) -> bool:
    return path.is_file() and (
        path.name == STATE_FILE_NAME or bool(LEGACY_ARTIFACT_NAME_PATTERN.match(path.name))
    )


def migrate_legacy_attempt(job_root: Path) -> None:
    if not job_root.exists():
        return

    legacy_files = [child for child in job_root.iterdir() if is_legacy_artifact(child)]
    if not legacy_files:
        return

    first_attempt_dir = attempt_dir(job_root, 1)
    if first_attempt_dir.exists():
        raise AutomationError(
            "Found both root-level generated files and attempt_1 under "
            f"{job_root}. Move the root-level files manually before continuing."
        )

    first_attempt_dir.mkdir(parents=True, exist_ok=True)
    for source in legacy_files:
        shutil.move(str(source), str(first_attempt_dir / source.name))
    log.info("Migrated %d root-level file(s) into %s", len(legacy_files), first_attempt_dir)


def parse_routine_ref(sql_text: str) -> RoutineRef:
    match = re.search(r"\bCALL\s+`([^`]+)`\s*\(", sql_text, flags=re.IGNORECASE)
    if not match:
        match = re.search(r"\bCALL\s+([A-Za-z0-9_.-]+)\s*\(", sql_text, flags=re.IGNORECASE)
    if not match:
        raise AutomationError("Could not find CALL routine in 2_sp_details.sql")

    parts = match.group(1).split(".")
    if len(parts) != 3:
        raise AutomationError(f"Expected project.dataset.routine in CALL, got: {match.group(1)}")
    project, dataset, routine_name = parts
    for part_name, part_value in [
        ("project", project),
        ("dataset", dataset),
        ("routine_name", routine_name),
    ]:
        if not re.fullmatch(r"[A-Za-z0-9_-]+", part_value):
            raise AutomationError(f"Invalid {part_name} in routine reference: {part_value}")
    return RoutineRef(project=project, dataset=dataset, routine_name=routine_name)


def parse_create_routine_ref(sql_text: str) -> RoutineRef:
    match = re.search(
        r"(?is)\bCREATE\s+(?:OR\s+REPLACE\s+)?PROCEDURE\s+"
        r"(?P<name>`[^`]+`|[A-Za-z0-9_.:-]+)",
        sql_text,
    )
    if not match:
        raise AutomationError("Could not find CREATE PROCEDURE routine in SP SQL")
    return parse_routine_name_text(match.group("name"))


def parse_routine_name_text(name_text: str) -> RoutineRef:
    parts = name_text.strip().strip("`").split(".")
    if len(parts) != 3:
        raise AutomationError(f"Expected project.dataset.routine, got: {name_text}")
    project, dataset, routine_name = parts
    for part_name, part_value in [
        ("project", project),
        ("dataset", dataset),
        ("routine_name", routine_name),
    ]:
        if not re.fullmatch(r"[A-Za-z0-9_-]+", part_value):
            raise AutomationError(f"Invalid {part_name} in routine reference: {part_value}")
    return RoutineRef(project=project, dataset=dataset, routine_name=routine_name)


def opt_prefixed_name(name: str) -> str:
    if name.casefold().startswith("opt_"):
        return name
    return f"opt_{name}"


def build_create_test_sp_sql(optimized_sp: str, test_routine: RoutineRef) -> str:
    sql, _ = build_create_test_sp_components(optimized_sp, test_routine)
    return sql


def build_create_test_sp_components(
    optimized_sp: str,
    test_routine: RoutineRef,
) -> tuple[str, list[DmlTargetMapping]]:
    procedure_sql = replace_create_procedure_routine(
        normalize_newlines(optimized_sp).strip(),
        test_routine,
    )
    procedure_sql, mappings = rewrite_prod_dml_targets(procedure_sql, test_routine)
    lines = [
        "-- ---------------------------------------------------------------------------",
        "-- Test scaffolding only.",
        f"-- The tables below are created in {test_routine.project}.{test_routine.dataset}",
        "-- as prerequisites for manually testing the optimized SP.",
        "-- They are not part of the stored procedure definition and are dropped in",
        "-- the cleanup block at the end of the combined test script.",
        "-- ---------------------------------------------------------------------------",
        "",
    ]
    if mappings:
        for mapping in mappings:
            test_table_name = format_table_name(
                test_routine.project,
                test_routine.dataset,
                mapping.temp_name,
            )
            lines.extend(
                [
                    f"CREATE OR REPLACE TABLE {test_table_name}",
                    f"LIKE `{mapping.ref.sql_name}`;",
                    "",
                ]
            )
    else:
        lines.extend(["-- No prod DML targets were detected.", ""])

    lines.append(procedure_sql.rstrip())
    return "\n".join(lines).rstrip() + "\n", mappings


def build_test_sp_sql(
    optimized_sp: str,
    sp_details: str,
    test_routine: RoutineRef,
    job_id: str | None = None,
) -> str:
    create_sql, mappings = build_create_test_sp_components(optimized_sp, test_routine)
    invoke_sql = rewrite_call_routine(normalize_newlines(sp_details).strip(), test_routine)
    lines = [
        "-- ---------------------------------------------------------------------------",
        "-- Test stored procedure script.",
        f"-- Creates scratch objects in {test_routine.project}.{test_routine.dataset},",
        "-- invokes the optimized test SP, and drops those objects at the end.",
        "-- ---------------------------------------------------------------------------",
        "",
        "-- ---------------------------------------------------------------------------",
        "-- 1. Create scratch tables and optimized test stored procedure.",
        "-- ---------------------------------------------------------------------------",
        "",
        create_sql.rstrip(),
        "",
        "-- ---------------------------------------------------------------------------",
        "-- 2. Invoke optimized test stored procedure.",
        "-- ---------------------------------------------------------------------------",
        "",
        "BEGIN",
        indent_sql(invoke_sql.rstrip()),
        "END;",
        "",
        "-- ---------------------------------------------------------------------------",
        "-- 3. Cleanup scratch tables and optimized test stored procedure.",
        "-- ---------------------------------------------------------------------------",
        "",
        *build_test_sp_cleanup_lines(test_routine, mappings),
    ]
    if job_id:
        lines = [f"-- Job ID: {single_line_sql_comment_text(job_id)}", "", *lines]
    return "\n".join(lines).rstrip() + "\n"


def job_id_from_artifact_dir(job_dir: Path) -> str:
    attempt_number = parse_attempt_dir_number(job_dir.name)
    if attempt_number is not None:
        return job_dir.parent.name
    return job_dir.name


def single_line_sql_comment_text(value: str) -> str:
    return " ".join(str(value).splitlines())


def build_invoke_test_sp_sql(
    sp_details: str,
    test_routine: RoutineRef,
    mappings: list[DmlTargetMapping] | None = None,
) -> str:
    invoke_sql = rewrite_call_routine(normalize_newlines(sp_details).strip(), test_routine)
    cleanup_lines = build_test_sp_cleanup_lines(test_routine, mappings or [])
    return f"{invoke_sql.rstrip()}\n\n" + "\n".join(cleanup_lines) + "\n"


def build_test_sp_cleanup_lines(
    test_routine: RoutineRef,
    mappings: list[DmlTargetMapping],
) -> list[str]:
    cleanup_lines = [
        "DROP TABLE IF EXISTS "
        f"{format_table_name(test_routine.project, test_routine.dataset, mapping.temp_name)};"
        for mapping in mappings
    ]
    cleanup_lines.append(f"DROP PROCEDURE IF EXISTS {format_routine_name(test_routine)};")
    return cleanup_lines


def indent_sql(sql_text: str, prefix: str = "  ") -> str:
    return "\n".join(f"{prefix}{line}" if line else "" for line in sql_text.splitlines())


def replace_create_procedure_routine(sql_text: str, test_routine: RoutineRef) -> str:
    pattern = re.compile(
        r"(?is)\bCREATE\s+(?:OR\s+REPLACE\s+)?PROCEDURE\s+"
        r"(?:`[^`]+`|[A-Za-z0-9_.:-]+)"
    )
    match = pattern.search(sql_text)
    if not match:
        raise AutomationError("Could not find CREATE PROCEDURE routine in SP SQL")
    replacement = f"CREATE OR REPLACE PROCEDURE {format_routine_name(test_routine)}"
    return sql_text[: match.start()] + replacement + sql_text[match.end() :]


def rewrite_call_routine(sql_text: str, test_routine: RoutineRef) -> str:
    pattern = re.compile(
        r"(?is)\bCALL\s+(?P<name>`[^`]+`|[A-Za-z0-9_.:-]+)\s*\("
    )
    match = pattern.search(sql_text)
    if not match:
        raise AutomationError("Could not find CALL routine in SP details SQL")
    start, end = match.span("name")
    return sql_text[:start] + format_routine_name(test_routine) + sql_text[end:]


def format_routine_name(routine: RoutineRef) -> str:
    name = f"{routine.project}.{routine.dataset}.{routine.routine_name}"
    return format_sql_object_name(name)


def format_table_name(project: str, dataset: str, table: str) -> str:
    return format_sql_object_name(f"{project}.{dataset}.{table}")


def format_sql_object_name(name: str) -> str:
    if re.fullmatch(
        r"[A-Za-z_][A-Za-z0-9_]*\.[A-Za-z_][A-Za-z0-9_]*\.[A-Za-z_][A-Za-z0-9_]*",
        name,
    ):
        return name
    return f"`{name}`"


def rewrite_prod_dml_targets(
    sql_text: str,
    test_routine: RoutineRef,
) -> tuple[str, list[DmlTargetMapping]]:
    occurrences = collect_prod_dml_target_occurrences(sql_text)
    mappings = assign_temp_table_names(occurrences)
    mapping_by_key = {mapping.ref.key: mapping for mapping in mappings}
    mapped_occurrences = collect_mapped_table_ref_occurrences(sql_text, mapping_by_key)
    rewritten = sql_text
    for occurrence in sorted(mapped_occurrences, key=lambda item: item.start, reverse=True):
        mapping = mapping_by_key[occurrence.ref.key]
        rewritten = (
            rewritten[: occurrence.start]
            + format_mapping_replacement(sql_text, occurrence, mapping, test_routine)
            + rewritten[occurrence.end :]
        )
    return rewritten, mappings


def format_mapping_replacement(
    sql_text: str,
    occurrence: DmlTargetOccurrence,
    mapping: DmlTargetMapping,
    test_routine: RoutineRef,
) -> str:
    raw_name = f"{test_routine.project}.{test_routine.dataset}.{mapping.temp_name}"
    if is_backticked_span(sql_text, occurrence.start, occurrence.end):
        return raw_name
    return format_sql_object_name(raw_name)


def is_backticked_span(sql_text: str, start: int, end: int) -> bool:
    return start > 0 and end < len(sql_text) and sql_text[start - 1] == "`" and sql_text[end] == "`"


def collect_prod_dml_target_mappings(sql_text: str) -> list[DmlTargetMapping]:
    return assign_temp_table_names(collect_prod_dml_target_occurrences(sql_text))


def collect_prod_dml_target_occurrences(sql_text: str) -> list[DmlTargetOccurrence]:
    visible = sql_visible_text(sql_text)
    table_ref_pattern = r"[A-Za-z0-9_-]+(?::|\.)[A-Za-z0-9_-]+\.[A-Za-z0-9_-]+"
    pattern = re.compile(
        rf"(?is)\b(?:"
        rf"INSERT\s+(?:INTO\s+)?|"
        rf"UPDATE\s+|"
        rf"DELETE\s+FROM\s+|"
        rf"MERGE\s+(?:INTO\s+)?|"
        rf"TRUNCATE\s+(?:TABLE\s+)?"
        rf")(?P<target>{table_ref_pattern})"
    )

    occurrences: list[DmlTargetOccurrence] = []
    for match in pattern.finditer(visible):
        start, end = match.span("target")
        ref = parse_table_ref(sql_text[start:end])
        if not ref or not is_prod_project(ref.project):
            continue
        occurrences.append(
            DmlTargetOccurrence(
                ref=ref,
                start=start,
                end=end,
                order=len(occurrences),
            )
        )
    return occurrences


def collect_mapped_table_ref_occurrences(
    sql_text: str,
    mapping_by_key: dict[tuple[str, str, str], DmlTargetMapping],
) -> list[DmlTargetOccurrence]:
    if not mapping_by_key:
        return []

    visible = sql_visible_text(sql_text)
    table_ref_pattern = re.compile(
        r"(?is)\b(?P<target>[A-Za-z0-9_-]+(?::|\.)[A-Za-z0-9_-]+\.[A-Za-z0-9_-]+)"
    )
    occurrences: list[DmlTargetOccurrence] = []
    for match in table_ref_pattern.finditer(visible):
        start, end = match.span("target")
        ref = parse_table_ref(sql_text[start:end])
        if not ref or ref.key not in mapping_by_key:
            continue
        occurrences.append(
            DmlTargetOccurrence(
                ref=ref,
                start=start,
                end=end,
                order=len(occurrences),
            )
        )
    return occurrences


def assign_temp_table_names(
    occurrences: list[DmlTargetOccurrence],
) -> list[DmlTargetMapping]:
    mappings: dict[tuple[str, str, str], DmlTargetMapping] = {}
    used_names: dict[str, tuple[str, str, str]] = {}
    for occurrence in occurrences:
        key = occurrence.ref.key
        if key in mappings:
            continue

        base_name = f"opt_{sanitize_sql_identifier(occurrence.ref.table)}"
        temp_name = base_name
        temp_key = temp_name.casefold()
        if temp_key in used_names and used_names[temp_key] != key:
            temp_name = (
                f"opt_{sanitize_sql_identifier(occurrence.ref.dataset)}_"
                f"{sanitize_sql_identifier(occurrence.ref.table)}"
            )
            temp_key = temp_name.casefold()

        suffix = 2
        while temp_key in used_names and used_names[temp_key] != key:
            temp_name = f"{base_name}_{suffix}"
            temp_key = temp_name.casefold()
            suffix += 1

        mapping = DmlTargetMapping(
            ref=occurrence.ref,
            temp_name=temp_name,
            order=occurrence.order,
        )
        mappings[key] = mapping
        used_names[temp_key] = key

    return sorted(mappings.values(), key=lambda item: item.order)


def parse_table_ref(name_text: str) -> TableRef | None:
    text = name_text.strip().strip("`")
    if ":" in text:
        project, remainder = text.split(":", 1)
        parts = remainder.split(".")
        if len(parts) != 2:
            return None
        dataset, table = parts
    else:
        parts = text.split(".")
        if len(parts) != 3:
            return None
        project, dataset, table = parts

    if not all([project, dataset, table]):
        return None
    return TableRef(project=project, dataset=dataset, table=table, raw=text)


def is_prod_project(project: str) -> bool:
    return "prod" in project.casefold()


def sanitize_sql_identifier(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", value.strip())
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    if not cleaned:
        cleaned = "table"
    if not re.match(r"[A-Za-z_]", cleaned):
        cleaned = f"_{cleaned}"
    return cleaned


SQL_CONTEXT_KEYWORDS = {
    "ALL",
    "AND",
    "AS",
    "ASC",
    "BEGIN",
    "BETWEEN",
    "BY",
    "CASE",
    "CREATE",
    "CURRENT_DATE",
    "CURRENT_DATETIME",
    "DATE",
    "DECLARE",
    "DEFAULT",
    "DESC",
    "DISTINCT",
    "DO",
    "ELSE",
    "END",
    "EXCEPT",
    "FALSE",
    "FROM",
    "GROUP",
    "HAVING",
    "IF",
    "IN",
    "INNER",
    "INSERT",
    "INTERVAL",
    "INTO",
    "IS",
    "JOIN",
    "LAST_DAY",
    "LEFT",
    "LIMIT",
    "MERGE",
    "MONTH",
    "NOT",
    "NULL",
    "ON",
    "OR",
    "ORDER",
    "OUT",
    "QUALIFY",
    "REPLACE",
    "RIGHT",
    "SELECT",
    "SET",
    "TABLE",
    "TEMP",
    "TEMPORARY",
    "THEN",
    "TRUE",
    "UNION",
    "UPDATE",
    "USING",
    "WHEN",
    "WHERE",
    "WHILE",
    "WITH",
}


def build_exec_query_context(
    sp_details: str,
    original_sp: str,
    original_query_text: str,
    optimized_query_text: str,
) -> str:
    sp_text = normalize_newlines(original_sp)
    statements = split_sql_statements(sp_text)
    target_index, target_span = find_original_query_statement(
        sp_text,
        statements,
        original_query_text,
    )
    warnings: list[str] = []
    if target_index is None:
        warnings.append(
            "Could not locate the original query inside 3_orig_sp.sql; "
            "stored procedure context could not be inferred."
        )
        prior_statements: list[SqlStatement] = []
    else:
        prior_statements = statements[:target_index]

    declarations = collect_declarations(prior_statements)
    set_statements = collect_set_statements(prior_statements)
    prior_temp_tables = collect_temp_table_statements(prior_statements)

    original_query = strip_query_destination_safe(original_query_text)
    optimized_query = strip_query_destination_safe(optimized_query_text)
    needed_temp_keys = resolve_needed_temp_tables(
        [original_query, optimized_query],
        prior_temp_tables,
    )
    context_temp_tables = [
        temp_table
        for temp_table in prior_temp_tables
        if temp_table.key in needed_temp_keys
    ]

    context_body_sql = "\n".join(
        [temp_table.text for temp_table in context_temp_tables]
        + [original_query, optimized_query]
    )
    temp_keys = {temp_table.key for temp_table in prior_temp_tables}
    needed_variable_keys = collect_declared_variable_refs(
        context_body_sql,
        declarations,
        temp_keys,
    )
    needed_variable_keys = expand_variable_dependencies(
        needed_variable_keys,
        declarations,
        set_statements,
        temp_keys,
    )

    declaration_lines = ordered_declarations_for_variables(
        declarations,
        needed_variable_keys,
    )
    set_lines = ordered_set_statements_for_variables(
        set_statements,
        needed_variable_keys,
    )
    parameter_lines, parameter_warnings = build_needed_parameter_declarations(
        sp_details,
        original_sp,
        "\n".join([context_body_sql] + set_lines),
    )
    warnings.extend(parameter_warnings)

    warnings.extend(
        find_unresolved_context_warnings(
            [context_body_sql, "\n".join(set_lines)],
            declarations,
            temp_keys,
            needed_temp_keys,
            sp_text[: target_span[0]] if target_span else "",
        )
    )
    loop_detected = bool(target_span and re.search(r"(?is)\bWHILE\b", sp_text[: target_span[0]]))

    return format_exec_context_block(
        parameter_lines,
        declaration_lines,
        set_lines,
        [temp_table.text for temp_table in context_temp_tables],
        warnings,
        loop_detected,
    )


def split_sql_statements(sql_text: str) -> list[SqlStatement]:
    text = normalize_newlines(sql_text)
    statements: list[SqlStatement] = []
    start = 0
    i = 0
    state: str | None = None
    quote = ""

    while i < len(text):
        if state == "line_comment":
            if text[i] == "\n":
                state = None
            i += 1
            continue
        if state == "block_comment":
            if text.startswith("*/", i):
                state = None
                i += 2
            else:
                i += 1
            continue
        if state == "single":
            if text[i] == "\\":
                i += 2
            elif text.startswith("''", i):
                i += 2
            elif text[i] == "'":
                state = None
                i += 1
            else:
                i += 1
            continue
        if state == "double":
            if text[i] == "\\":
                i += 2
            elif text.startswith('""', i):
                i += 2
            elif text[i] == '"':
                state = None
                i += 1
            else:
                i += 1
            continue
        if state == "triple":
            if text.startswith(quote, i):
                state = None
                i += 3
            else:
                i += 1
            continue
        if state == "backtick":
            if text[i] == "`":
                state = None
            i += 1
            continue

        if text.startswith("--", i):
            state = "line_comment"
            i += 2
            continue
        if text[i] == "#":
            state = "line_comment"
            i += 1
            continue
        if text.startswith("/*", i):
            state = "block_comment"
            i += 2
            continue
        if text.startswith("'''", i) or text.startswith('"""', i):
            quote = text[i : i + 3]
            state = "triple"
            i += 3
            continue
        if text[i] == "'":
            state = "single"
            i += 1
            continue
        if text[i] == '"':
            state = "double"
            i += 1
            continue
        if text[i] == "`":
            state = "backtick"
            i += 1
            continue
        if text[i] == ";":
            statement_text = text[start : i + 1]
            if statement_text.strip():
                statements.append(SqlStatement(statement_text, start, i + 1))
            start = i + 1
        i += 1

    trailing = text[start:]
    if trailing.strip():
        statements.append(SqlStatement(trailing, start, len(text)))
    return statements


def sql_visible_text(sql_text: str) -> str:
    text = normalize_newlines(sql_text)
    chars: list[str] = []
    i = 0
    state: str | None = None
    quote = ""

    while i < len(text):
        if state == "line_comment":
            chars.append("\n" if text[i] == "\n" else " ")
            if text[i] == "\n":
                state = None
            i += 1
            continue
        if state == "block_comment":
            if text.startswith("*/", i):
                chars.extend("  ")
                state = None
                i += 2
            else:
                chars.append("\n" if text[i] == "\n" else " ")
                i += 1
            continue
        if state in {"single", "double"}:
            escaped_pair = "''" if state == "single" else '""'
            closing_quote = "'" if state == "single" else '"'
            if text[i] == "\\":
                chars.extend("  "[: max(0, min(2, len(text) - i))])
                i += 2
            elif text.startswith(escaped_pair, i):
                chars.extend("  ")
                i += 2
            elif text[i] == closing_quote:
                chars.append(" ")
                state = None
                i += 1
            else:
                chars.append("\n" if text[i] == "\n" else " ")
                i += 1
            continue
        if state == "triple":
            if text.startswith(quote, i):
                chars.extend("   ")
                state = None
                i += 3
            else:
                chars.append("\n" if text[i] == "\n" else " ")
                i += 1
            continue
        if state == "backtick":
            if text[i] == "`":
                chars.append(" ")
                state = None
            else:
                chars.append(text[i])
            i += 1
            continue

        if text.startswith("--", i):
            chars.extend("  ")
            state = "line_comment"
            i += 2
            continue
        if text[i] == "#":
            chars.append(" ")
            state = "line_comment"
            i += 1
            continue
        if text.startswith("/*", i):
            chars.extend("  ")
            state = "block_comment"
            i += 2
            continue
        if text.startswith("'''", i) or text.startswith('"""', i):
            quote = text[i : i + 3]
            chars.extend("   ")
            state = "triple"
            i += 3
            continue
        if text[i] == "'":
            chars.append(" ")
            state = "single"
            i += 1
            continue
        if text[i] == '"':
            chars.append(" ")
            state = "double"
            i += 1
            continue
        if text[i] == "`":
            chars.append(" ")
            state = "backtick"
            i += 1
            continue

        chars.append(text[i])
        i += 1

    return "".join(chars)


def find_original_query_statement(
    original_sp: str,
    statements: list[SqlStatement],
    original_query_text: str,
) -> tuple[int | None, tuple[int, int] | None]:
    query_span = find_query_span(original_sp, original_query_text)
    if query_span is None:
        return None, None
    for index, statement in enumerate(statements):
        if statement.start <= query_span[0] < statement.end:
            return index, query_span
    return None, query_span


def find_query_span(base_sql: str, query_text: str) -> tuple[int, int] | None:
    base = normalize_newlines(base_sql)
    query = normalize_newlines(query_text).strip()
    candidates = [query, query.rstrip(";").rstrip() + ";", query.rstrip(";").rstrip()]
    try:
        stripped = strip_query_destination(query)
        candidates.extend(
            [
                stripped,
                stripped.rstrip(";").rstrip() + ";",
                stripped.rstrip(";").rstrip(),
            ]
        )
    except AutomationError:
        pass

    for candidate in unique_preserving_order(candidates):
        if not candidate:
            continue
        index = base.find(candidate)
        if index >= 0:
            return index, index + len(candidate)
    return None


def collect_declarations(
    statements: list[SqlStatement],
) -> dict[str, tuple[int, str]]:
    declarations: dict[str, tuple[int, str]] = {}
    for order, statement in enumerate(statements):
        parsed = parse_declare_statement(statement.text)
        if not parsed:
            continue
        names, text = parsed
        for name in names:
            declarations.setdefault(normalize_identifier_key(name), (order, text))
    return declarations


def parse_declare_statement(statement_text: str) -> tuple[list[str], str] | None:
    visible = sql_visible_text(statement_text)
    match = re.search(
        r"(?is)\bDECLARE\s+"
        r"([A-Za-z_][A-Za-z0-9_]*(?:\s*,\s*[A-Za-z_][A-Za-z0-9_]*)*)\s+",
        visible,
    )
    if not match:
        return None
    names = [name.strip() for name in match.group(1).split(",")]
    return names, ensure_semicolon(statement_text[match.start() :].strip())


def collect_set_statements(
    statements: list[SqlStatement],
) -> dict[str, list[tuple[int, str]]]:
    set_statements: dict[str, list[tuple[int, str]]] = {}
    for order, statement in enumerate(statements):
        parsed = parse_set_statement(statement.text)
        if not parsed:
            continue
        name, text = parsed
        set_statements.setdefault(normalize_identifier_key(name), []).append((order, text))
    return set_statements


def parse_set_statement(statement_text: str) -> tuple[str, str] | None:
    visible = sql_visible_text(statement_text)
    match = re.search(r"(?is)\bSET\s+([A-Za-z_][A-Za-z0-9_]*)\s*=", visible)
    if not match:
        return None
    return match.group(1), ensure_semicolon(statement_text[match.start() :].strip())


def collect_temp_table_statements(statements: list[SqlStatement]) -> list[TempTableStatement]:
    temp_tables: list[TempTableStatement] = []
    for order, statement in enumerate(statements):
        parsed = parse_temp_table_statement(statement.text)
        if not parsed:
            continue
        name, text = parsed
        temp_tables.append(
            TempTableStatement(
                name=name,
                key=normalize_identifier_key(simple_identifier_name(name)),
                text=text,
                order=order,
            )
        )
    return temp_tables


def parse_temp_table_statement(statement_text: str) -> tuple[str, str] | None:
    visible = sql_visible_text(statement_text)
    match = re.search(
        r"(?is)\bCREATE\s+(?:OR\s+REPLACE\s+)?"
        r"(?P<temp>TEMP(?:ORARY)?\s+)?TABLE\s+"
        r"(?:IF\s+NOT\s+EXISTS\s+)?"
        r"(?P<name>[A-Za-z_][A-Za-z0-9_.:-]*)",
        visible,
    )
    if not match or not match.group("temp"):
        return None
    return match.group("name"), ensure_semicolon(statement_text[match.start() :].strip())


def resolve_needed_temp_tables(
    seed_sql: list[str],
    prior_temp_tables: list[TempTableStatement],
) -> set[str]:
    needed = find_referenced_temp_table_keys(seed_sql, prior_temp_tables)
    changed = True
    while changed:
        changed = False
        for temp_table in prior_temp_tables:
            if temp_table.key not in needed:
                continue
            dependencies = find_referenced_temp_table_keys(
                [temp_table.text],
                prior_temp_tables,
            )
            for dependency in dependencies:
                if dependency not in needed:
                    needed.add(dependency)
                    changed = True
    return needed


def find_referenced_temp_table_keys(
    sql_parts: list[str],
    prior_temp_tables: list[TempTableStatement],
) -> set[str]:
    found: set[str] = set()
    joined = "\n".join(sql_parts)
    for temp_table in prior_temp_tables:
        if references_identifier(joined, temp_table.name):
            found.add(temp_table.key)
    return found


def collect_declared_variable_refs(
    sql_text: str,
    declarations: dict[str, tuple[int, str]],
    temp_keys: set[str],
) -> set[str]:
    refs: set[str] = set()
    for variable_key in declarations:
        if variable_key in temp_keys:
            continue
        if references_identifier(sql_text, variable_key):
            refs.add(variable_key)
    return refs


def expand_variable_dependencies(
    needed_variable_keys: set[str],
    declarations: dict[str, tuple[int, str]],
    set_statements: dict[str, list[tuple[int, str]]],
    temp_keys: set[str],
) -> set[str]:
    expanded = set(needed_variable_keys)
    changed = True
    while changed:
        changed = False
        for variable_key in list(expanded):
            for _, set_text in set_statements.get(variable_key, []):
                rhs = extract_set_rhs(set_text)
                for dependency in collect_declared_variable_refs(rhs, declarations, temp_keys):
                    if dependency not in expanded:
                        expanded.add(dependency)
                        changed = True
    return expanded


def extract_set_rhs(set_text: str) -> str:
    visible = sql_visible_text(set_text)
    match = re.search(r"(?is)\bSET\s+[A-Za-z_][A-Za-z0-9_]*\s*=\s*", visible)
    if not match:
        return set_text
    return set_text[match.end() :].strip().rstrip(";")


def ordered_declarations_for_variables(
    declarations: dict[str, tuple[int, str]],
    needed_variable_keys: set[str],
) -> list[str]:
    selected: list[tuple[int, str]] = []
    for variable_key in needed_variable_keys:
        declaration = declarations.get(variable_key)
        if declaration:
            selected.append(declaration)
    return unique_ordered_text(selected)


def ordered_set_statements_for_variables(
    set_statements: dict[str, list[tuple[int, str]]],
    needed_variable_keys: set[str],
) -> list[str]:
    selected: list[tuple[int, str]] = []
    for variable_key in needed_variable_keys:
        selected.extend(set_statements.get(variable_key, []))
    return unique_ordered_text(selected)


def unique_ordered_text(items: list[tuple[int, str]]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for _, text in sorted(items, key=lambda item: item[0]):
        if text not in seen:
            seen.add(text)
            result.append(text)
    return result


def build_needed_parameter_declarations(
    sp_details: str,
    original_sp: str,
    context_sql: str,
) -> tuple[list[str], list[str]]:
    warnings: list[str] = []
    params = parse_procedure_params(original_sp)
    if not params:
        return [], warnings
    call_args = parse_call_args(sp_details)
    if not call_args:
        return [], ["Could not find a CALL statement in 2_sp_details.sql."]

    input_params = [
        param
        for param in params
        if param.mode.upper() in {"IN", "INOUT"}
    ]
    needed_param_keys = {
        normalize_identifier_key(param.name)
        for param in input_params
        if references_identifier(context_sql, param.name)
    }

    lines: list[str] = []
    for index, param in enumerate(params):
        param_key = normalize_identifier_key(param.name)
        if param.mode.upper() not in {"IN", "INOUT"} or param_key not in needed_param_keys:
            continue
        if index >= len(call_args):
            warnings.append(f"CALL argument for procedure parameter {param.name} was not found.")
            continue
        arg = call_args[index].strip().rstrip(";")
        if not arg:
            warnings.append(f"CALL argument for procedure parameter {param.name} is empty.")
            continue
        lines.append(f"DECLARE {param.name} {param.type_sql} DEFAULT {arg};")
    return lines, warnings


def parse_procedure_params(original_sp: str) -> list[ProcedureParam]:
    text = normalize_newlines(original_sp)
    match = re.search(
        r"(?is)\bCREATE\s+(?:OR\s+REPLACE\s+)?PROCEDURE\s+"
        r"(?:`[^`]+`|[A-Za-z0-9_.:-]+)\s*\(",
        text,
    )
    if not match:
        return []
    open_index = match.end() - 1
    close_index = find_matching_paren(text, open_index)
    if close_index is None:
        return []

    params: list[ProcedureParam] = []
    for raw_param in split_top_level_commas(text[open_index + 1 : close_index]):
        param = raw_param.strip()
        if not param:
            continue
        param_match = re.match(
            r"(?is)^(?:(INOUT|IN|OUT)\s+)?([A-Za-z_][A-Za-z0-9_]*)\s+(.+)$",
            param,
        )
        if not param_match:
            continue
        mode = (param_match.group(1) or "IN").upper()
        params.append(
            ProcedureParam(
                name=param_match.group(2),
                mode=mode,
                type_sql=param_match.group(3).strip(),
            )
        )
    return params


def parse_call_args(sp_details: str) -> list[str]:
    text = normalize_newlines(sp_details)
    match = re.search(
        r"(?is)\bCALL\s+(?:`[^`]+`|[A-Za-z0-9_.:-]+)\s*\(",
        text,
    )
    if not match:
        return []
    open_index = match.end() - 1
    close_index = find_matching_paren(text, open_index)
    if close_index is None:
        return []
    return split_top_level_commas(text[open_index + 1 : close_index])


def find_matching_paren(text: str, open_index: int) -> int | None:
    visible = sql_visible_text(text)
    depth = 0
    for index in range(open_index, len(visible)):
        if visible[index] == "(":
            depth += 1
        elif visible[index] == ")":
            depth -= 1
            if depth == 0:
                return index
    return None


def split_top_level_commas(text: str) -> list[str]:
    visible = sql_visible_text(text)
    parts: list[str] = []
    start = 0
    paren_depth = 0
    bracket_depth = 0
    angle_depth = 0
    for index, char in enumerate(visible):
        if char == "(":
            paren_depth += 1
        elif char == ")" and paren_depth:
            paren_depth -= 1
        elif char == "[":
            bracket_depth += 1
        elif char == "]" and bracket_depth:
            bracket_depth -= 1
        elif char == "<":
            angle_depth += 1
        elif char == ">" and angle_depth:
            angle_depth -= 1
        elif char == "," and not paren_depth and not bracket_depth and not angle_depth:
            parts.append(text[start:index].strip())
            start = index + 1
    parts.append(text[start:].strip())
    return parts


def find_unresolved_context_warnings(
    sql_parts: list[str],
    declarations: dict[str, tuple[int, str]],
    temp_keys: set[str],
    needed_temp_keys: set[str],
    prior_sql: str,
) -> list[str]:
    warnings: list[str] = []
    joined = "\n".join(sql_parts)
    declared_keys = set(declarations)
    for variable_name in sorted(find_likely_variable_references(joined)):
        variable_key = normalize_identifier_key(variable_name)
        if variable_key not in declared_keys and variable_key not in temp_keys:
            warnings.append(
                f"Verify variable {variable_name}; no prior DECLARE was found in the SP."
            )

    local_refs = find_likely_local_table_references(joined)
    for table_name in sorted(local_refs):
        table_key = normalize_identifier_key(simple_identifier_name(table_name))
        if table_key in {"V_TEMP_TABLE_ORIG", "V_TEMP_TABLE_OPT"}:
            continue
        if table_key not in needed_temp_keys and table_key not in temp_keys:
            detail = f"Verify temp table {table_name}; no prior temp-table creation was found."
            if re.search(r"(?is)\bEXECUTE\s+IMMEDIATE\b", prior_sql):
                detail += " It may be created by dynamic SQL."
            warnings.append(detail)

    return unique_preserving_order(warnings)


def find_likely_variable_references(sql_text: str) -> set[str]:
    visible = sql_visible_text(sql_text)
    refs: set[str] = set()
    for match in re.finditer(r"\b[A-Za-z_][A-Za-z0-9_]*\b", visible):
        name = match.group(0)
        key = normalize_identifier_key(name)
        if key in SQL_CONTEXT_KEYWORDS:
            continue
        if key.startswith("V_"):
            refs.add(name)
    return refs


def find_likely_local_table_references(sql_text: str) -> set[str]:
    visible = sql_visible_text(sql_text)
    refs: set[str] = set()
    pattern = re.compile(
        r"(?is)\b(?:FROM|JOIN|TABLE|INTO|UPDATE|MERGE\s+INTO)\s+"
        r"([A-Za-z_][A-Za-z0-9_.:-]*)"
    )
    for match in pattern.finditer(visible):
        name = match.group(1)
        if "." in name:
            continue
        if is_likely_temp_identifier(name):
            refs.add(name)
    return refs


def is_likely_temp_identifier(name: str) -> bool:
    key = normalize_identifier_key(simple_identifier_name(name))
    return (
        key.startswith("V_")
        or key.startswith("TMP")
        or key.startswith("TEMP")
        or key.endswith("_TMP")
        or key.endswith("_TEMP")
        or "_TMP_" in key
        or "_TEMP_" in key
    )


def format_exec_context_block(
    parameter_lines: list[str],
    declaration_lines: list[str],
    set_lines: list[str],
    temp_table_lines: list[str],
    warnings: list[str],
    loop_detected: bool,
) -> str:
    lines = [
        "-- START STORED PROCEDURE CONTEXT",
        "-- Auto-generated from 2_sp_details.sql and 3_orig_sp.sql.",
    ]
    if loop_detected:
        lines.append(
            "-- Loop context: this uses the first/default loop state; "
            "edit the DECLARE/SET values below to test another iteration."
        )
    if warnings:
        lines.append("-- WARNING: Review the TODO items before relying on this validation script.")
        for warning in warnings:
            lines.append(f"-- TODO: {warning}")

    if not any([parameter_lines, declaration_lines, set_lines, temp_table_lines]):
        lines.append("-- No stored procedure context dependencies were detected.")
        lines.append("-- END STORED PROCEDURE CONTEXT")
        return "\n".join(lines) + "\n"

    append_context_section(lines, parameter_lines)
    append_context_section(lines, declaration_lines)
    append_context_section(lines, set_lines)
    append_context_section(lines, temp_table_lines)
    lines.append("-- END STORED PROCEDURE CONTEXT")
    return "\n".join(lines) + "\n"


def append_context_section(lines: list[str], section_lines: list[str]) -> None:
    if not section_lines:
        return
    if lines and lines[-1] != "":
        lines.append("")
    lines.extend(section_lines)


def references_identifier(sql_text: str, identifier: str) -> bool:
    key = normalize_identifier_key(simple_identifier_name(identifier))
    visible = sql_visible_text(sql_text)
    pattern = re.compile(rf"(?i)(?<![A-Za-z0-9_]){re.escape(key)}(?![A-Za-z0-9_])")
    return bool(pattern.search(visible.upper()))


def normalize_identifier_key(identifier: str) -> str:
    return simple_identifier_name(identifier).strip("`").upper()


def simple_identifier_name(identifier: str) -> str:
    text = identifier.strip().strip("`")
    if "." in text:
        return text.split(".")[-1].strip("`")
    return text


def strip_query_destination_safe(sql: str) -> str:
    try:
        return strip_query_destination(sql)
    except AutomationError:
        return normalize_newlines(sql).strip()


def strip_query_destination(sql: str) -> str:
    text = normalize_newlines(sql).strip()
    while True:
        stripped = strip_leading_create_table(strip_leading_insert(text))
        if stripped == text:
            return stripped
        text = stripped


def strip_leading_insert(sql: str) -> str:
    text = normalize_newlines(sql).strip()
    pattern = re.compile(
        r"(?is)^insert\s+into\s+"
        r"(?:`[^`]+`|[A-Za-z0-9_.:-]+)"
        r"(?:\s*\([^)]*\))?"
        r"\s+",
    )
    match = pattern.match(text)
    if not match:
        if text.lower().startswith("insert"):
            raise AutomationError("Query starts with INSERT but the target table could not be parsed")
        return text
    return text[match.end() :].strip()


def strip_leading_create_table(sql: str) -> str:
    text = normalize_newlines(sql).strip()
    pattern = re.compile(
        r"(?is)^create\s+(?:or\s+replace\s+)?(?:temp(?:orary)?\s+)?table\s+"
        r"(?:if\s+not\s+exists\s+)?"
        r"(?:`[^`]+`|[A-Za-z0-9_.:-]+)"
        r"(?:\s*\([^)]*\))?"
        r"\s+as\s+",
    )
    match = pattern.match(text)
    if not match:
        return text
    return text[match.end() :].strip()


def ensure_semicolon(sql: str) -> str:
    return normalize_newlines(sql).strip().rstrip(";").rstrip() + ";"


def replace_original_query(original_sp: str, original_query: str, replacement: str) -> str:
    base = normalize_newlines(original_sp)
    query = normalize_newlines(original_query).strip()
    candidates = unique_preserving_order(
        [
            query,
            query.rstrip(";").rstrip() + ";",
            query.rstrip(";").rstrip(),
        ]
    )
    for candidate in candidates:
        if candidate and candidate in base:
            return base.replace(candidate, replacement, 1)
    raise AutomationError(
        "Could not find the original query text inside 3_orig_sp.sql. "
        "The SP file was left unchanged."
    )


def unique_preserving_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result


def find_duplicates(values: list[str]) -> set[str]:
    seen: set[str] = set()
    duplicates: set[str] = set()
    for value in values:
        if value in seen:
            duplicates.add(value)
        seen.add(value)
    return duplicates


def get_optimized_created_at(state: dict[str, Any]) -> str:
    created_at = state.get("optimized_result", {}).get("created_at")
    if not created_at:
        raise AutomationError("optimized_result.created_at is missing from checkpoint state")
    return str(created_at)


def stringify_timestamp(value: Any) -> str:
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def coerce_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)

    text = str(value).strip()
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError as exc:
        raise AutomationError(f"Expected DATETIME-compatible value, got: {value!r}") from exc
    return parsed.replace(tzinfo=None)


def make_authorized_session() -> AuthorizedSession:
    credentials, _ = google.auth.default(
        scopes=["https://www.googleapis.com/auth/cloud-platform"]
    )
    return AuthorizedSession(credentials)


def escape_sql_string(value: str) -> str:
    return value.replace("\\", "\\\\").replace("'", "\\'")


def escape_triple_quoted_string(value: str) -> str:
    return value.replace('"""', '\\"\\"\\"')


def normalize_newlines(value: str) -> str:
    return value.replace("\r\n", "\n").replace("\r", "\n")


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def read_text_if_exists(path: Path) -> str:
    if not path.exists():
        return ""
    return read_text(path)


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(normalize_newlines(text), encoding="utf-8", newline="\n")
    
def fetch_jobs_by_owner(csv_path: str, target_owner: str) -> List[dict]:
    """Fetch deliverables from the CSV based on 'owner' AND 'Status'."""
    jobs = []
    target_status = "in progress"
    
    with open(csv_path, mode='r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        # enumerate(..., start=2) tracks the actual Excel/CSV row number
        for row_num, row in enumerate(reader, start=2):
            owner = row.get('owner', '').strip().lower()
            status = row.get('Status', '').strip().lower()
            
            if owner == target_owner.lower() and status == target_status:
                # Store the entire row dictionary and inject our row number
                job_data = dict(row)
                job_data['csv_row_number'] = row_num
                jobs.append(job_data)
                
    return jobs

def process_single_job(row: dict, base_args: argparse.Namespace, force_rerun_ids: List[str], progress: Progress, task_id: int) -> Dict[str, str]:
    """Handles artifact generation for a single job ID."""
    
    job_id = row.get('job_id', '').strip()
    progress.update(task_id, description=f"[cyan]Processing {job_id} - Initializing...")
    
    try:
        item = ConfigItem(
            entity=row.get('Entity', '').strip(), 
            metric_name=row.get('metric_name', '').strip(), 
            parent_job_id=row.get('parent_job_id', '').strip(),
            job_id=job_id
        )
        
        local_args = copy.copy(base_args)
        
        if job_id in force_rerun_ids:
            local_args.rerun_job_id = job_id  
            progress.update(task_id, description=f"[yellow]Processing {job_id} - FORCING NEW WORKFLOW...")
        else:
            local_args.rerun_job_id = None
            progress.update(task_id, description=f"[blue]Processing {job_id} - Fetching/Generating Artifacts...")
            
        automation = ArtifactAutomation(local_args)
        
        # --- GENERATE DOCUMENTATION FIRST ---
        attempt_number = automation.resolve_attempt_number(item)
        job_dir = item.job_root / f"attempt_{attempt_number}"
        
        # Create the directory structure manually first so we can write the doc
        job_dir.mkdir(parents=True, exist_ok=True)
        
        doc_file = job_dir / "0_documentation.md"
        doc_content = (
            f"# Job Documentation: {job_id}\n\n"
            f"- **CSV Row Number:** {row.get('csv_row_number')}\n"
            f"- **Owner:** {row.get('owner', 'N/A')}\n"
            f"- **Status:** {row.get('Status', 'N/A')}\n"
            f"- **Entity:** {row.get('Entity', 'N/A')}\n"
            f"- **Job ID:** {job_id}\n"
            f"- **Parent Job ID:** {row.get('parent_job_id', 'N/A')}\n"
            f"- **SP Name:** {row.get('SP_Name', 'N/A')}\n"
            f"- **Frequency:** {row.get('frequency', 'N/A')}\n\n"
            f"## Notes\n{row.get('notes', 'None provided.')}\n\n"
            f"## Comments\n{row.get('comments', 'None provided.')}\n\n"
            f"## Error Message\n{row.get('error_message', 'None provided.')}\n"
        )
        doc_file.write_text(doc_content, encoding="utf-8")
        # ----------------------------------------------

        # Now execute the heavy BigQuery operations
        automation.process_item(item)
        
        progress.update(task_id, description=f"[green]Completed {job_id}", completed=100)
        
        status_msg = "SUCCESS: Forced new workflow execution." if job_id in force_rerun_ids else "SUCCESS: Artifacts built and fetched."
        return {"job_id": job_id, "status": status_msg}

    except Exception as e:
        progress.update(task_id, description=f"[red]Failed {job_id}", completed=100)
        return {"job_id": job_id, "status": f"ERROR: {str(e)}"}

def run_concurrent_batch(csv_path: str, owner: str, args: argparse.Namespace, force_rerun_ids: List[str], max_workers: int = 5, preview_only: bool = False):
    """Orchestrates the concurrent execution and CLI dashboard."""
    target_jobs = fetch_jobs_by_owner(csv_path, owner)
    
    if not target_jobs:
        console.print(f"[bold red]No jobs found for owner: '{owner}'[/bold red]")
        return
    
    if preview_only:
        console.print(f"\n[bold cyan]🔍 PREVIEW MODE: Found {len(target_jobs)} jobs. No actions will be taken.[/bold cyan]")
        table = Table(show_header=True, header_style="bold magenta")
        table.add_column("Row #", justify="right", style="dim")
        table.add_column("Job ID", style="cyan")
        table.add_column("Entity", style="green")
        table.add_column("Status", style="blue")
        table.add_column("Action", style="yellow")
        
        for row in target_jobs:
            job_id = row.get('job_id', 'Unknown')
            # Highlight if the job is going to force a brand new workflow run
            action = "[bold red]FORCED RERUN[/bold red]" if job_id in force_rerun_ids else "Standard (Fetch/Gen)"
            
            table.add_row(
                str(row.get('csv_row_number', 'N/A')),
                job_id,
                row.get('Entity', 'N/A'),
                row.get('Status', 'N/A'),
                action
            )
        console.print(table)
        console.print("[dim]Run the command again without --preview to execute this batch.[/dim]\n")
        return

    console.print(f"[bold green]Found {len(target_jobs)} jobs for {owner}. Starting concurrent processing with {max_workers} workers...[/bold green]")
    if force_rerun_ids:
        console.print(f"[bold yellow]Forcing new workflow execution for {len(force_rerun_ids)} flagged jobs.[/bold yellow]")

    results = []
    
    # 2 refreshes per second prevents the Git Bash "waterfall" rendering glitch
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
        console=console,
        refresh_per_second=2, 
    ) as progress:
        
        futures_map = {}
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            for row in target_jobs:
                job_id = row.get('job_id', 'Unknown')
                task_id = progress.add_task(f"[cyan]Pending {job_id}", total=100)
                
                # Submit to thread pool
                future = executor.submit(process_single_job, row, args, force_rerun_ids, progress, task_id)
                futures_map[future] = job_id
            
            for future in concurrent.futures.as_completed(futures_map):
                job_id = futures_map[future]
                try:
                    result = future.result()
                    results.append(result)
                except Exception as exc:
                    results.append({"job_id": job_id, "status": f"FATAL THREAD ERROR: {exc}"})

    console.print("\n[bold]Execution Summary:[/bold]")
    for res in results:
        status_color = "green" if "SUCCESS" in res['status'] else "red"
        console.print(f"Job: [bold]{res['job_id']}[/bold] | Status: [{status_color}]{res['status']}[/{status_color}]")


if __name__ == "__main__":
    # 1. Pre-parse custom flags before passing the rest to your original parser
    custom_parser = argparse.ArgumentParser(add_help=False)
    custom_parser.add_argument("--force-rerun-ids", type=str, default="", help="Comma-separated list of job IDs to force rerun")
    
    # ADDITIONAL ARGUMENTS
    custom_parser.add_argument("--owner", type=str, default="Ajman", help="Target owner name (Defaults to 'Ajman')")
    custom_parser.add_argument("--csv", type=str, default="config.csv", help="Path to your target CSV file")
    custom_parser.add_argument("--preview", action="store_true", help="Print a table of jobs to be executed without actually running them")
    
    custom_args, remaining_argv = custom_parser.parse_known_args(sys.argv[1:])
    
    # 2. Parse standard arguments using your existing parser
    args = parse_args(remaining_argv)
    
    # 3. Clean up the comma-separated force-rerun list
    force_ids_list = [jid.strip() for jid in custom_args.force_rerun_ids.split(",")] if custom_args.force_rerun_ids else []
    
    try:
        run_concurrent_batch(
            csv_path=custom_args.csv, 
            owner=custom_args.owner, 
            args=args,
            force_rerun_ids=force_ids_list,
            max_workers=5,
            preview_only=custom_args.preview  # Pass the preview flag
        )
    except Exception as e:
        console.print(f"[bold red]Execution halted: {e}[/bold red]")